# -*- coding: utf-8 -*-
"""Funções utilitárias para captura do GTFS"""

import io
import json
import os
import re
from pathlib import Path
from typing import Union

import openpyxl as xl
import pandas as pd
import requests
from google.auth import default
from google.cloud import bigquery, storage
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from unidecode import unidecode

from pipelines.capture__gtfs import constants
from pipelines.common import constants as smtr_constants
from pipelines.common.utils.gcp.bigquery import BQTable, Dataset
from pipelines.common.utils.utils import is_running_locally


def get_google_api_service(service_name: str, version: str, scopes=None):
    """Retorna um serviço do Google API configurado com as credenciais apropriadas."""
    scopes_by_service = {
        "sheets": ["https://www.googleapis.com/auth/spreadsheets.readonly"],
        "drive": ["https://www.googleapis.com/auth/drive.readonly"],
    }
    if scopes is None:
        scopes = scopes_by_service.get(
            service_name, ["https://www.googleapis.com/auth/drive.readonly"]
        )

    if is_running_locally():
        creds, _ = default(scopes=scopes)
    else:
        creds = service_account.Credentials.from_service_account_file(
            filename=os.environ["GOOGLE_APPLICATION_CREDENTIALS"], scopes=scopes
        )

    return build(service_name, version, credentials=creds)


def get_upload_storage_blob(env: str, dataset_id: str, filename: str):
    """Retorna um blob da zona de upload do GCS."""
    client = storage.Client()
    bucket_name = smtr_constants.DEFAULT_BUCKET_NAME[env]
    bucket = client.bucket(bucket_name)
    blobs = list(bucket.list_blobs(prefix=f"upload/{dataset_id}/{filename}."))
    if not blobs:
        raise FileNotFoundError(f"Nenhum blob encontrado em upload/{dataset_id}/{filename}")
    return blobs[0]


def xl_load_workbook_sheetnames(file_bytes: io.BytesIO) -> list:
    """Retorna os nomes das abas de um arquivo Excel."""
    file_bytes.seek(0)
    wb = xl.load_workbook(file_bytes)
    names = wb.sheetnames
    file_bytes.seek(0)
    return names


def save_raw_local_func(
    data: Union[dict, str],
    filepath: str,
    mode: str = "raw",
    filetype: str = "json",
) -> str:
    """Salva dados brutos em um arquivo local."""
    _filepath = filepath.format(mode=mode, filetype=filetype)
    Path(_filepath).parent.mkdir(parents=True, exist_ok=True)

    if filetype == "json":
        if isinstance(data, str):
            data = json.loads(data)
        with Path(_filepath).open("w", encoding="utf-8") as fi:
            json.dump(data, fi)

    elif filetype in ("txt", "csv"):
        with Path(_filepath).open("w", encoding="utf-8") as file:
            file.write(data)

    print(f"Raw data saved to: {_filepath}")
    return _filepath


def create_bq_external_table(env: str, dataset_id: str, table_id: str, staging_filepath: str):
    """Cria uma tabela externa no BigQuery apontando para o GCS."""
    staging_dataset_id = f"{dataset_id}_staging"
    tb_obj = BQTable(env=env, dataset_id=staging_dataset_id, table_id=table_id)

    Dataset(dataset_id=staging_dataset_id, env=env).create()

    bq_client = tb_obj.client("bigquery")
    bucket_name = tb_obj.bucket_name
    source_uri_prefix = f"gs://{bucket_name}/staging/{dataset_id}/{table_id}/"

    sample_df = pd.read_csv(staging_filepath, nrows=1)
    schema = [bigquery.SchemaField(col, "STRING") for col in sample_df.columns]

    external_config = bigquery.ExternalConfig("CSV")
    external_config.source_uris = [f"{source_uri_prefix}*/*"]
    external_config.options.skip_leading_rows = 1
    external_config.schema = schema

    hive_opts = bigquery.HivePartitioningOptions()
    hive_opts.mode = "AUTO"
    hive_opts.source_uri_prefix = source_uri_prefix
    external_config.hive_partitioning = hive_opts

    bq_table = bigquery.Table(tb_obj.table_full_name)
    prod_table_name = tb_obj.table_full_name.replace(staging_dataset_id, dataset_id, 1)
    bq_table.description = f"staging table for `{prod_table_name}`"
    bq_table.external_data_configuration = external_config

    bq_client.create_table(bq_table)
    print(f"Tabela externa {tb_obj.table_full_name} criada com sucesso.")


def filter_valid_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Filtra linhas válidas do DataFrame de controle de OS."""
    df["index"] = df.index
    df.dropna(how="all", inplace=True)
    df = df[df["Fim da Vigência da OS"] != "Sem Vigência"]
    df = df[df["Submeter mudanças para Dados"] == True]  # noqa
    df = df[~df["Início da Vigência da OS"].isnull()]
    df = df[~df["Arquivo OS"].isnull()]
    df = df[~df["Arquivo GTFS"].isnull()]
    df = df[~df["Link da OS"].isnull()]
    df = df[~df["Link do GTFS"].isnull()]
    return df


def download_controle_os_csv(url: str) -> pd.DataFrame:
    """Baixa o CSV de controle de OS e retorna como DataFrame."""
    response = requests.get(url=url, timeout=smtr_constants.MAX_TIMEOUT_SECONDS)
    response.raise_for_status()
    response.encoding = "utf-8"
    df = pd.read_csv(io.StringIO(response.text))
    print(f"Download concluído! Dados:\n{df.head()}")
    return df


def convert_to_float(value):
    """Converte um valor string para float."""
    if "," in str(value):
        value = str(value).replace(".", "").replace(",", ".").strip()
    return float(value)


def normalizar_horario(horario: str) -> str:
    """Normaliza uma string de horário."""
    horario = str(horario)
    if "day" in horario:
        days, time = horario.split(", ")
        days = int(days.split(" ")[0])
        hours, minutes, seconds = map(int, time.split(":"))
        total_hours = days * 24 + hours
        return f"{total_hours:02}:{minutes:02}:{seconds:02}"
    else:
        return horario.split(" ")[1] if " " in horario else horario


def download_xlsx(file_link: str, drive_service) -> io.BytesIO:
    """Baixa um arquivo XLSX do Google Drive."""
    file_id = file_link.split("/")[-2]
    file = drive_service.files().get(fileId=file_id, supportsAllDrives=True).execute()
    mime_type = file.get("mimeType")

    if "google-apps" in mime_type:
        request = drive_service.files().export(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        request = drive_service.files().get_media(fileId=file_id)

    file_bytes = io.BytesIO()
    downloader = MediaIoBaseDownload(file_bytes, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    file_bytes.seek(0)
    return file_bytes


def download_file(file_link: str, drive_service) -> io.BytesIO:
    """Baixa um arquivo do Google Drive."""
    file_id = file_link.split("/")[-2]
    request = drive_service.files().get_media(fileId=file_id, supportsAllDrives=True)
    file_bytes = io.BytesIO()
    downloader = MediaIoBaseDownload(file_bytes, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return file_bytes


def processa_ordem_servico(
    sheetnames,
    file_bytes,
    local_filepath,
    raw_filepaths,
    regular_sheet_index=None,  # noqa: ARG001
):
    """Processa as abas de Ordem de Serviço de um arquivo Excel."""
    sheets = [(i, name) for i, name in enumerate(sheetnames) if "ANEXO I " in name]
    if not sheets:
        raise ValueError("Nenhuma aba 'ANEXO I' encontrada no arquivo.")
    sheets_data = []

    columns = {
        "Serviço": "servico",
        "Vista": "vista",
        "Consórcio": "consorcio",
        "Extensão de Ida": "extensao_ida",
        "Extensão de Volta": "extensao_volta",
        "Horário Inicial": "horario_inicio",
        "Horário\nInicial": "horario_inicio",
        "Horário Fim": "horario_fim",
        "Horário\nFim": "horario_fim",
        "Partidas Ida Dia Útil": "partidas_ida_du",
        "Partidas Volta Dia Útil": "partidas_volta_du",
        "Viagens Dia Útil": "viagens_du",
        "Quilometragem Dia Útil": "km_dia_util",
        "KM Dia Útil": "km_dia_util",
        "Partidas Ida Sábado": "partidas_ida_sabado",
        "Partidas Volta Sábado": "partidas_volta_sabado",
        "Viagens Sábado": "viagens_sabado",
        "Quilometragem Sábado": "km_sabado",
        "KM Sábado": "km_sabado",
        "Partidas Ida Domingo": "partidas_ida_domingo",
        "Partidas Volta Domingo": "partidas_volta_domingo",
        "Viagens Domingo": "viagens_domingo",
        "Quilometragem Domingo": "km_domingo",
        "KM Domingo": "km_domingo",
        "Partidas Ida Ponto Facultativo": "partidas_ida_pf",
        "Partidas Volta Ponto Facultativo": "partidas_volta_pf",
        "Viagens Ponto Facultativo": "viagens_pf",
        "Quilometragem Ponto Facultativo": "km_pf",
        "KM Ponto Facultativo": "km_pf",
        "tipo_os": "tipo_os",
    }

    for _, sheet_name in sheets:
        print(f"########## {sheet_name} ##########")
        match = re.search(r"\((.*?)\)", sheet_name)
        if not match:
            raise ValueError(f"Não foi possível extrair tipo_os do nome da aba: {sheet_name}")
        tipo_os = match.group(1)

        quadro = pd.read_excel(file_bytes, sheet_name=sheet_name, dtype=object)
        quadro = quadro.rename(columns=columns)
        quadro["servico"] = quadro["servico"].astype(str)
        quadro["servico"] = quadro["servico"].str.extract(r"([A-Z]+)", expand=False).fillna(
            ""
        ) + quadro["servico"].str.extract(r"([0-9]+)", expand=False).fillna("")
        quadro["tipo_os"] = tipo_os
        quadro = quadro[list(set(columns.values()))]
        quadro = quadro.replace("—", 0)
        quadro = quadro.reindex(columns=list(set(columns.values())))

        hora_cols = [coluna for coluna in quadro.columns if "horario" in coluna]
        quadro[hora_cols] = quadro[hora_cols].astype(str)
        for hora_col in hora_cols:
            quadro[hora_col] = quadro[hora_col].apply(normalizar_horario)

        cols = [
            coluna
            for coluna in quadro.columns
            if "km" in coluna or "viagens" in coluna or "partida" in coluna
        ]
        for col in cols:
            quadro[col] = quadro[col].astype(str).apply(convert_to_float).astype(float).fillna(0)

        extensao_cols = ["extensao_ida", "extensao_volta"]
        quadro[extensao_cols] = quadro[extensao_cols].astype(str)
        for col in extensao_cols:
            quadro[col] = quadro[col].str.replace(".", "", regex=False)
        quadro[extensao_cols] = quadro[extensao_cols].apply(pd.to_numeric)
        quadro["extensao_ida"] = quadro["extensao_ida"] / 1000
        quadro["extensao_volta"] = quadro["extensao_volta"] / 1000

        sheets_data.append(quadro)

    quadro_geral = pd.concat(sheets_data, ignore_index=True)

    columns_in_dataframe = set(quadro_geral.columns)
    columns_in_values = set(columns.values())
    all_columns_present = columns_in_dataframe.issubset(columns_in_values)
    no_duplicate_columns = len(columns_in_dataframe) == len(quadro_geral.columns)
    missing_columns = columns_in_values - columns_in_dataframe

    print(
        f"All columns present: {all_columns_present}/"
        f"No duplicate columns: {no_duplicate_columns}/"
        f"Missing columns: {missing_columns}"
    )

    if not all_columns_present or not no_duplicate_columns:
        raise Exception("Missing or duplicated columns in ordem_servico")

    local_file_path = next(filter(lambda x: "ordem_servico/" in x, local_filepath))
    quadro_geral_csv = quadro_geral.to_csv(index=False)
    raw_file_path = save_raw_local_func(
        data=quadro_geral_csv, filepath=local_file_path, filetype="csv"
    )
    print(f"Saved file: {raw_file_path}")
    raw_filepaths.append(raw_file_path)


def processa_ordem_servico_trajeto_alternativo(
    sheetnames,
    file_bytes,
    local_filepath,
    raw_filepaths,
    data_versao_gtfs,
    filename,
):
    """Processa as abas de Trajetos Alternativos de um arquivo Excel."""
    sheets = [(i, name) for i, name in enumerate(sheetnames) if "ANEXO II " in name]
    if not sheets:
        raise ValueError("Nenhuma aba 'ANEXO II' encontrada no arquivo.")
    sheets_data = []

    if data_versao_gtfs < constants.DATA_GTFS_V5_INICIO:
        alt_columns = {
            "Serviço": "servico",
            "Vista": "vista",
            "Consórcio": "consorcio",
            "Extensão de Ida": "extensao_ida",
            "Extensão\nde Ida": "extensao_ida",
            "Extensão de Volta": "extensao_volta",
            "Extensão\nde Volta": "extensao_volta",
            "Evento": "evento",
            "Horário Inicial Interdição": "inicio_periodo",
            "Horário Final Interdição": "fim_periodo",
            "Descrição": "descricao",
            "Ativação": "ativacao",
            "tipo_os": "tipo_os",
        }
    else:
        alt_columns = {
            "Serviço": "servico",
            "Vista": "vista",
            "Sentido": "sentido",
            "Extensão": "extensao",
            "Consórcio": "consorcio",
            "Evento": "evento",
            "Descrição": "descricao",
            "Ativação": "ativacao",
            "tipo_os": "tipo_os",
        }

    for _, sheet_name in sheets:
        print(f"########## {sheet_name} ##########")
        match = re.search(r"\((.*?)\)", sheet_name)
        if not match:
            raise ValueError(f"Não foi possível extrair tipo_os do nome da aba: {sheet_name}")
        tipo_os = match.group(1)

        df = pd.read_excel(file_bytes, sheet_name=sheet_name, dtype=object)
        df = df.rename(columns=alt_columns)
        df["tipo_os"] = tipo_os
        sheets_data.append(df)

    ordem_servico_trajeto_alternativo = pd.concat(sheets_data, ignore_index=True)
    columns_in_dataframe = set(ordem_servico_trajeto_alternativo.columns)
    columns_in_values = set(alt_columns.values())
    all_columns_present = columns_in_dataframe.issubset(columns_in_values)
    no_duplicate_columns = len(columns_in_dataframe) == len(
        ordem_servico_trajeto_alternativo.columns
    )
    missing_columns = columns_in_values - columns_in_dataframe

    print(
        f"All columns present: {all_columns_present}/"
        f"No duplicate columns: {no_duplicate_columns}/"
        f"Missing columns: {missing_columns}"
    )

    if not all_columns_present or not no_duplicate_columns:
        raise Exception("Missing or duplicated columns in ordem_servico_trajeto_alternativo")

    local_file_path = next(filter(lambda x: filename + "/" in x, local_filepath))
    csv_data = ordem_servico_trajeto_alternativo.to_csv(index=False)
    raw_file_path = save_raw_local_func(data=csv_data, filepath=local_file_path, filetype="csv")
    print(f"Saved file: {raw_file_path}")
    raw_filepaths.append(raw_file_path)


def processa_ordem_servico_faixa_horaria(  # noqa: PLR0912, PLR0915
    sheetnames,
    file_bytes,
    local_filepath,
    raw_filepaths,
    data_versao_gtfs,
    filename="ordem_servico_faixa_horaria",
):
    """Processa as abas de Faixa Horária de um arquivo Excel."""
    if data_versao_gtfs >= constants.DATA_GTFS_V2_INICIO:
        sheets = [(i, name) for i, name in enumerate(sheetnames) if "ANEXO I " in name]
        if not sheets:
            raise ValueError("Nenhuma aba 'ANEXO I' encontrada no arquivo.")
    else:
        sheets = [(i, name) for i, name in enumerate(sheetnames) if "ANEXO III " in name]
        if not sheets:
            raise ValueError("Nenhuma aba 'ANEXO III' encontrada no arquivo.")
    sheets_data = []

    columns = {
        "Serviço": "servico",
        "Vista": "vista",
        "Consórcio": "consorcio",
        "Extensão de Ida": "extensao_ida",
        "Extensão de Volta": "extensao_volta",
        "Horário Inicial Dias Úteis": "horario_inicio_dias_uteis",
        "Horário Fim Dias Úteis": "horario_fim_dias_uteis",
        "Horário Inicial - Dias Úteis": "horario_inicio_dias_uteis",
        "Horário Fim - Dias Úteis": "horario_fim_dias_uteis",
        "Partidas Ida - Dias Úteis": "partidas_ida_dias_uteis",
        "Partidas Volta - Dias Úteis": "partidas_volta_dias_uteis",
        "Viagens - Dias Úteis": "viagens_dias_uteis",
        "Quilometragem - Dias Úteis": "quilometragem_dias_uteis",
        "KM - Dias Úteis": "quilometragem_dias_uteis",
        "Horário Inicial Sábado": "horario_inicio_sabado",
        "Horário Fim Sábado": "horario_fim_sabado",
        "Horário Inicial - Sábado": "horario_inicio_sabado",
        "Horário Fim - Sábado": "horario_fim_sabado",
        "Partidas Ida - Sábado": "partidas_ida_sabado",
        "Partidas Volta - Sábado": "partidas_volta_sabado",
        "Viagens - Sábado": "viagens_sabado",
        "Quilometragem - Sábado": "quilometragem_sabado",
        "KM - Sábado": "quilometragem_sabado",
        "Horário Inicial Domingo": "horario_inicio_domingo",
        "Horário Fim Domingo": "horario_fim_domingo",
        "Horário Inicial - Domingo": "horario_inicio_domingo",
        "Horário Fim - Domingo": "horario_fim_domingo",
        "Partidas Ida - Domingo": "partidas_ida_domingo",
        "Partidas Volta - Domingo": "partidas_volta_domingo",
        "Viagens - Domingo": "viagens_domingo",
        "Quilometragem - Domingo": "quilometragem_domingo",
        "KM - Domingo": "quilometragem_domingo",
        "Horário Inicial Ponto Facultativo": "horario_inicio_ponto_facultativo",
        "Horário Fim Ponto Facultativo": "horario_fim_ponto_facultativo",
        "Horário Inicial - Ponto Facultativo": "horario_inicio_ponto_facultativo",
        "Horário Fim - Ponto Facultativo": "horario_fim_ponto_facultativo",
        "Partidas Ida - Ponto Facultativo": "partidas_ida_ponto_facultativo",
        "Partidas Volta - Ponto Facultativo": "partidas_volta_ponto_facultativo",
        "Viagens - Ponto Facultativo": "viagens_ponto_facultativo",
        "Quilometragem - Ponto Facultativo": "quilometragem_ponto_facultativo",
        "KM - Ponto Facultativo": "quilometragem_ponto_facultativo",
        "tipo_os": "tipo_os",
    }

    metricas = ["Partidas", "Partidas Ida", "Partidas Volta", "Quilometragem", "KM"]
    dias = ["Dias Úteis", "Sábado", "Domingo", "Ponto Facultativo"]
    formatos = [
        "{metrica} entre {intervalo} — {dia}",
        "{metrica} entre {intervalo} - {dia}",
        "{metrica} entre {intervalo} ({dia})",
        "{metrica} {intervalo} - {dia}",
    ]

    if data_versao_gtfs < constants.DATA_GTFS_V3_INICIO:
        intervalos = [
            "00h e 03h",
            "03h e 12h",
            "12h e 21h",
            "21h e 24h",
            "24h e 03h (dia seguinte)",
        ]
    elif (
        data_versao_gtfs >= constants.DATA_GTFS_V3_INICIO
        and data_versao_gtfs < constants.DATA_GTFS_V4_INICIO
    ):
        intervalos = [
            "00h e 03h",
            "03h e 06h",
            "06h e 09h",
            "09h e 12h",
            "12h e 15h",
            "15h e 18h",
            "18h e 21h",
            "21h e 24h",
            "24h e 03h (dia seguinte)",
        ]
    else:
        intervalos = [
            "00h à 01h",
            "01h à 02h",
            "02h à 03h",
            "03h à 04h",
            "04h à 05h",
            "05h à 06h",
            "06h à 09h",
            "09h à 12h",
            "12h à 15h",
            "15h à 18h",
            "18h à 21h",
            "21h à 22h",
            "22h à 23h",
            "23h à 24h",
        ]

    fh_columns = {
        formato.format(metrica=metrica, intervalo=intervalo, dia=dia): unidecode(
            (
                "quilometragem"
                if metrica in ["Quilometragem", "KM"]
                else (
                    "partidas"
                    if (metrica == "Partidas" and data_versao_gtfs < constants.DATA_GTFS_V2_INICIO)
                    or data_versao_gtfs >= constants.DATA_GTFS_V4_INICIO
                    else ("partidas_ida" if metrica == "Partidas Ida" else "partidas_volta")
                )
            )
            + f"_entre_{intervalo.replace(' ', '_').replace('(', '').replace(')', '').replace('-', '_')}_{dia.lower().replace(' ', '_')}"
        )
        for metrica in metricas
        for intervalo in intervalos
        for dia in dias
        for formato in formatos
    }

    if data_versao_gtfs >= constants.DATA_GTFS_V4_INICIO:
        fh_columns["Serviço"] = "servico"
        fh_columns["Consórcio"] = "consorcio"
        fh_columns["tipo_os"] = "tipo_os"
        fh_columns["Sentido"] = "sentido"
        fh_columns["Extensão"] = "extensao"
        fh_columns["Vista"] = "vista"
        columns = fh_columns.copy()
    elif data_versao_gtfs >= constants.DATA_GTFS_V2_INICIO:
        columns.update(fh_columns)
    else:
        fh_columns["Serviço"] = "servico"
        fh_columns["Consórcio"] = "consorcio"
        fh_columns["tipo_os"] = "tipo_os"
        columns = fh_columns.copy()

    for _, sheet_name in sheets:
        print(f"########## {sheet_name} ##########")
        match = re.search(r"\((.*?)\)", sheet_name)
        if not match:
            raise ValueError(f"Não foi possível extrair tipo_os do nome da aba: {sheet_name}")
        tipo_os = match.group(1)

        df = pd.read_excel(file_bytes, sheet_name=sheet_name, dtype=object)
        df.columns = (
            df.columns.str.replace("\n", " ").str.strip().str.replace(r"\s+", " ", regex=True)
        )
        df = df.rename(columns=lambda x: x.replace("Dia Útil", "Dias Úteis"))
        df = df.rename(columns=columns)

        for col in df.columns:
            if "quilometragem" in col or "viagens" in col or "partidas" in col:
                df[col] = df[col].astype(str).replace(["—", "-"], 0)
            if "quilometragem" in col or "viagens" in col:
                df[col] = df[col].astype(str).apply(convert_to_float).astype(float)
            if "extensao" in col:
                df[col] = df[col].apply(pd.to_numeric)
                df[col] = df[col] / 1000
            if "horario" in col:
                df[col] = df[col].astype(str)
                df[col] = df[col].apply(normalizar_horario)

        df["tipo_os"] = tipo_os

        columns_in_values = set(columns.values())
        aux_columns_in_dataframe = set(df.columns)
        aux_missing_columns = columns_in_values - aux_columns_in_dataframe
        for coluna in aux_missing_columns:
            if "ponto_facultativo" in coluna:
                df[coluna] = 0

        sheets_data.append(df)

    ordem_servico_faixa_horaria = pd.concat(sheets_data, ignore_index=True)
    columns_in_dataframe = set(ordem_servico_faixa_horaria.columns)
    columns_in_values = set(columns.values())
    missing_columns = columns_in_values - columns_in_dataframe
    all_columns_present = columns_in_dataframe.issubset(columns_in_values)
    no_duplicate_columns = len(columns_in_dataframe) == len(ordem_servico_faixa_horaria.columns)

    print(
        f"All columns present: {all_columns_present}\n"
        f"No duplicate columns: {no_duplicate_columns}\n"
        f"Missing columns: {missing_columns}"
    )

    if not all_columns_present or not no_duplicate_columns:
        print(columns_in_values.difference(columns_in_dataframe))
        print(columns_in_dataframe.difference(columns_in_values))
        raise Exception("Missing or duplicated columns in ordem_servico_faixa_horaria")

    local_file_path = next(filter(lambda x: filename + "/" in x, local_filepath))
    csv_data = ordem_servico_faixa_horaria.to_csv(index=False)
    raw_file_path = save_raw_local_func(data=csv_data, filepath=local_file_path, filetype="csv")
    print(f"Saved file: {raw_file_path}")
    raw_filepaths.append(raw_file_path)
